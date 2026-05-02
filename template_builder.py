"""ひな形 .xlsx を一から組み立てる。

`明細` シート (ヘッダー + 書式) と `集計` シート (月別/店舗別の SUMIFS) を持つ。
"""
from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from excel_writer import COLUMNS, DETAIL_SHEET

SUMMARY_SHEET = "集計"

HEADER_FILL = PatternFill("solid", fgColor="2563EB")
HEADER_FONT = Font(color="FFFFFF", bold=True)
NUMBER_FMT = "#,##0"
DATE_FMT = "yyyy/mm/dd"

COL_WIDTHS = {
    "A": 12, "B": 18, "C": 24, "D": 36, "E": 6, "F": 10, "G": 12,
    "H": 8, "I": 8, "J": 12, "K": 16, "L": 28, "M": 36, "N": 16, "O": 24,
}


def _setup_detail(ws):
    for col_idx, name in enumerate(COLUMNS, start=1):
        c = ws.cell(row=1, column=col_idx, value=name)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center", vertical="center")
    for letter, width in COL_WIDTHS.items():
        ws.column_dimensions[letter].width = width
    ws.freeze_panes = "A2"
    # 列ごとの数値書式（ヘッダ行以外）
    for col_letter, fmt in (
        ("A", DATE_FMT),
        ("E", "0"),
        ("F", NUMBER_FMT),
        ("G", NUMBER_FMT),
        ("H", NUMBER_FMT),
        ("I", NUMBER_FMT),
        ("J", NUMBER_FMT),
    ):
        for row in range(2, 1001):
            ws[f"{col_letter}{row}"].number_format = fmt
    # 行高
    ws.row_dimensions[1].height = 22


def _setup_summary(ws):
    ws["A1"] = "店舗別合計"
    ws["A1"].font = HEADER_FONT
    ws["A1"].fill = HEADER_FILL
    ws["A2"] = "店舗・サイト"
    ws["B2"] = "合計金額"
    ws["A2"].font = Font(bold=True)
    ws["B2"].font = Font(bold=True)

    # 動的に拾う数式（明細B列のユニーク → SUMIFS）。シンプルに上位20店舗まで枠を確保
    for row in range(3, 23):
        ws[f"A{row}"] = ""  # ユーザーが手で書くか、後で拡張
        ws[f"B{row}"] = (
            f'=IF(A{row}="",0,SUMIFS(明細!J:J,明細!B:B,A{row}))'
        )
        ws[f"B{row}"].number_format = NUMBER_FMT

    ws["D1"] = "月別合計"
    ws["D1"].font = HEADER_FONT
    ws["D1"].fill = HEADER_FILL
    ws["D2"] = "年月"
    ws["E2"] = "合計金額"
    ws["D2"].font = Font(bold=True)
    ws["E2"].font = Font(bold=True)
    for row in range(3, 23):
        ws[f"D{row}"] = ""  # 例: 2026-05
        ws[f"E{row}"] = (
            f'=IF(D{row}="",0,SUMPRODUCT((TEXT(明細!A:A,"yyyy-mm")=D{row})*明細!J:J))'
        )
        ws[f"E{row}"].number_format = NUMBER_FMT

    ws["G1"] = "全体合計"
    ws["G1"].font = HEADER_FONT
    ws["G1"].fill = HEADER_FILL
    ws["G2"] = "件数"
    ws["H2"] = "=COUNTIF(明細!J:J,\">0\")"
    ws["G3"] = "合計金額"
    ws["H3"] = "=SUM(明細!J:J)"
    ws["H3"].number_format = NUMBER_FMT

    for letter, width in {"A": 18, "B": 14, "D": 12, "E": 14, "G": 12, "H": 14}.items():
        ws.column_dimensions[letter].width = width


def build() -> bytes:
    wb = Workbook()
    detail = wb.active
    detail.title = DETAIL_SHEET
    _setup_detail(detail)

    summary = wb.create_sheet(SUMMARY_SHEET)
    _setup_summary(summary)

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


if __name__ == "__main__":
    import sys
    out_path = sys.argv[1] if len(sys.argv) > 1 else "template/receipts_template.xlsx"
    with open(out_path, "wb") as f:
        f.write(build())
    print(f"wrote {out_path}")
