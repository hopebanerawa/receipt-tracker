"""ひな形 .xlsx の `明細` シート末尾に行を追加して、書き戻し用のbytesを返す。"""
from __future__ import annotations

import io
from datetime import datetime
from typing import TypedDict

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

DETAIL_SHEET = "明細"
TEMPLATE_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

HEADER_ROW = 1
FIRST_DATA_ROW = 2

# A..O
COLUMNS = [
    "取引日",          # A
    "店舗・サイト",    # B
    "注文ID",          # C
    "商品名",          # D
    "数量",            # E
    "単価",            # F
    "商品小計",        # G
    "送料",            # H
    "手数料",          # I
    "合計金額",        # J
    "決済方法",        # K
    "元ファイル名",    # L
    "Driveリンク",     # M
    "処理日時",        # N
    "備考",            # O
]


class AppendedRow(TypedDict):
    取引日: str | None
    店舗: str
    注文ID: str | None
    商品名: str
    数量: int
    単価: int | None
    商品小計: int
    送料: int
    手数料: int
    合計金額: int
    決済方法: str | None
    元ファイル名: str
    Driveリンク: str
    処理日時: str
    備考: str


def _next_empty_row(ws) -> int:
    row = FIRST_DATA_ROW
    while ws.cell(row=row, column=1).value not in (None, ""):
        row += 1
    return row


def _parse_date(s: str | None):
    if not s:
        return None
    try:
        return datetime.strptime(s, "%Y-%m-%d").date()
    except ValueError:
        return s  # そのまま文字列で書く


def append_receipt(
    wb_bytes: bytes,
    *,
    receipt: dict,
    file_name: str,
    drive_link: str,
) -> tuple[bytes, list[AppendedRow]]:
    """1領収書をひな形に追記し、(更新後bytes, 追加行のリスト) を返す。"""
    wb = load_workbook(io.BytesIO(wb_bytes))
    if DETAIL_SHEET not in wb.sheetnames:
        raise ValueError(f"テンプレに '{DETAIL_SHEET}' シートがありません")
    ws = wb[DETAIL_SHEET]

    items = receipt.get("items") or []
    if not items:
        # 明細が抽出できなかった場合も合計1行は記録
        items = [{"name": "(明細抽出失敗)", "qty": 1, "unit_price": None, "subtotal": receipt.get("total", 0)}]

    now_str = datetime.now().strftime("%Y-%m-%d %H:%M")
    appended: list[AppendedRow] = []
    start = _next_empty_row(ws)
    txn_date = _parse_date(receipt.get("transaction_date"))

    for idx, it in enumerate(items):
        is_first = idx == 0
        row = start + idx
        values = [
            txn_date,
            receipt.get("vendor") or "",
            receipt.get("order_id") if is_first else None,
            it.get("name") or "",
            it.get("qty") or 1,
            it.get("unit_price"),
            it.get("subtotal") or 0,
            (receipt.get("shipping") or 0) if is_first else 0,
            (receipt.get("fee") or 0) if is_first else 0,
            (receipt.get("total") or 0) if is_first else 0,
            receipt.get("payment_method") if is_first else None,
            file_name if is_first else "",
            drive_link if is_first else "",
            now_str if is_first else "",
            "",
        ]
        for col_idx, val in enumerate(values, start=1):
            ws.cell(row=row, column=col_idx, value=val)

        appended.append(
            AppendedRow(
                取引日=str(txn_date) if txn_date else None,
                店舗=receipt.get("vendor") or "",
                注文ID=receipt.get("order_id"),
                商品名=it.get("name") or "",
                数量=it.get("qty") or 1,
                単価=it.get("unit_price"),
                商品小計=it.get("subtotal") or 0,
                送料=(receipt.get("shipping") or 0) if is_first else 0,
                手数料=(receipt.get("fee") or 0) if is_first else 0,
                合計金額=(receipt.get("total") or 0) if is_first else 0,
                決済方法=receipt.get("payment_method") if is_first else None,
                元ファイル名=file_name if is_first else "",
                Driveリンク=drive_link if is_first else "",
                処理日時=now_str if is_first else "",
                備考="",
            )
        )

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue(), appended


def read_summary(wb_bytes: bytes) -> dict:
    """累計件数・累計金額をざっくり集計（明細の最初行のみが合計を持つ前提）。"""
    wb = load_workbook(io.BytesIO(wb_bytes), data_only=False)
    if DETAIL_SHEET not in wb.sheetnames:
        return {"rows": 0, "total_amount": 0, "receipts": 0}
    ws = wb[DETAIL_SHEET]
    rows = 0
    receipts = 0
    total = 0
    r = FIRST_DATA_ROW
    while ws.cell(row=r, column=1).value not in (None, ""):
        rows += 1
        v = ws.cell(row=r, column=10).value  # J列: 合計金額
        if isinstance(v, (int, float)) and v > 0:
            receipts += 1
            total += int(v)
        r += 1
    return {"rows": rows, "total_amount": total, "receipts": receipts}
